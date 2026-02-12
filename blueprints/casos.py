import os
import io
from datetime import datetime, timedelta, date
from flask import Blueprint, render_template, abort, request, flash, redirect, url_for, send_file
from flask_login import login_required, current_user
from sqlalchemy import case, or_, func
from models import db, Caso, Usuario, Rol, AuditoriaCaso, CatalogoEstablecimiento, CatalogoInstitucion, CatalogoRecinto, obtener_hora_chile, CasoGestion
from utils import check_password_change, registrar_log, enviar_aviso_asignacion, generar_acta_cierre_pdf, enviar_aviso_cierre, enviar_aviso_subrogancia, es_rut_valido, safe_int, enviar_reporte_estadistico_masivo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

casos_bp = Blueprint('casos', __name__, template_folder='../templates', url_prefix='/casos')

def clean(value):
    """Convierte '' / espacios a None (para guardar NULL en BD)."""
    if value is None:
        return None
    v = str(value).strip()
    return v if v else None

def clean_rut(value):
    """Normaliza RUT: sin puntos/espacios y con guion (12345678-9)."""
    v = clean(value)
    if not v:
        return None
    v = v.replace(".", "").replace(" ", "").upper()
    # Si viene con guion, lo respetamos; si no, lo armamos.
    if "-" in v:
        cuerpo, dv = v.split("-", 1)
    else:
        if len(v) < 2:
            return None
        cuerpo, dv = v[:-1], v[-1]
    cuerpo = "".join(ch for ch in cuerpo if ch.isdigit())
    dv = (dv[:1] or "").upper()
    return f"{cuerpo}-{dv}" if cuerpo and dv else None

def rut_excede_largo(rut_normalizado):
    """Protege contra Data too long. Formato máximo esperado: 12345678-9 (10 chars)."""
    return bool(rut_normalizado) and len(rut_normalizado) > 10

@casos_bp.before_request
@login_required
@check_password_change
def before_request():
    pass

@casos_bp.route('/')
def index():
    """
    Bandeja de Entrada + Dashboard Ejecutivo (Fase 3 Refinada).
    """
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '').strip()
    estado_filter = request.args.get('estado', '').strip()

    rol_nombre = current_user.rol.nombre
    titulo_vista = "Vista Global"

    # =========================================================
    # A. FILTROS DE SEGURIDAD (ROLES) + SUBROGANCIA
    # =========================================================
    filters = []

    # Datos para UI Subrogancia (solo referentes)
    candidatos_subrogancia = []
    subrogante_activo = None  # objeto Usuario (el que me está subrogando)

    if rol_nombre == 'Admin':
        pass

    elif rol_nombre in ['Referente', 'Visualizador']:
        ciclos_permitidos = []

        # 1) Ciclo propio (si tiene)
        if current_user.ciclo_asignado_id:
            ciclos_permitidos.append(current_user.ciclo_asignado_id)
            titulo_vista = f"Ciclo {current_user.ciclo_asignado.nombre}"
        else:
            titulo_vista = "Vista Global (Todos los Ciclos)"

        # 2) Ciclo subrogado (si soy subrogante de alguien)
        if current_user.subrogante_de and current_user.subrogante_de.ciclo_asignado_id:
            ciclo_sub = current_user.subrogante_de.ciclo_asignado_id
            if ciclo_sub not in ciclos_permitidos:
                ciclos_permitidos.append(ciclo_sub)

            # Título más informativo
            if current_user.ciclo_asignado_id:
                titulo_vista = f"Ciclo {current_user.ciclo_asignado.nombre} + Subrogancia ({current_user.subrogante_de.ciclo_asignado.nombre})"
            else:
                titulo_vista = f"Subrogancia ({current_user.subrogante_de.ciclo_asignado.nombre})"

        # Aplicar filtro por ciclos permitidos si corresponde
        if ciclos_permitidos:
            filters.append(Caso.ciclo_vital_id.in_(ciclos_permitidos))
        else:
            # Si no tiene ciclo propio ni subrogado:
            # si tu regla es que "sin ciclo" = global, no filtramos.
            # (mantiene la lógica actual)
            pass

        # --- UI Subrogancia: solo si soy Referente (no Visualizador) ---
        if rol_nombre == 'Referente':
            # a) Lista de candidatos (otros referentes activos, distintos a mí)
            candidatos_subrogancia = Usuario.query.join(Rol).filter(
                Rol.nombre == 'Referente',
                Usuario.activo == True,
                Usuario.id != current_user.id
            ).order_by(Usuario.nombre_completo).all()

            # b) Subrogante activo (quién me está subrogando a mí)
            # OJO: subrogantes_activos es dynamic (AppenderQuery), así que usamos first()
            subrogante_activo = None
            if getattr(current_user, "subrogantes_activos", None):
                try:
                    subrogante_activo = current_user.subrogantes_activos.first()
                except Exception:
                    # Si por alguna razón NO es query y sí es lista, igual funcionará
                    subrogante_activo = current_user.subrogantes_activos[0] if current_user.subrogantes_activos else None

    elif rol_nombre == 'Funcionario':
        filters.append(Caso.asignado_a_usuario_id == current_user.id)
        titulo_vista = "Mis Casos Asignados"

    else:
        abort(403)

    # Query Base (Solo Seguridad)
    base_query = Caso.query.filter(*filters)

    # =========================================================
    # B. DASHBOARD DATA (KPIs + GRÁFICOS)
    # Ignoran filtros de búsqueda/estado, muestran la "realidad total" del usuario
    # =========================================================
    
    # 1. KPIs Generales
    stats_query = db.session.query(
        func.count(Caso.id).label('total'),
        func.sum(case((Caso.estado == 'PENDIENTE_RESCATAR', 1), else_=0)).label('pendientes'),
        func.sum(case((Caso.estado == 'EN_SEGUIMIENTO', 1), else_=0)).label('seguimiento'),
        func.sum(case((Caso.estado == 'CERRADO', 1), else_=0)).label('cerrados')
    ).filter(*filters)
    
    stats_result = stats_query.first()
    
    total = stats_result.total or 0
    pendientes = int(stats_result.pendientes or 0)
    seguimiento = int(stats_result.seguimiento or 0)
    cerrados = int(stats_result.cerrados or 0)

    # Calculamos porcentajes para la vista (evitar división por cero)
    pct_pendientes = round((pendientes / total * 100), 1) if total > 0 else 0
    pct_seguimiento = round((seguimiento / total * 100), 1) if total > 0 else 0
    pct_cerrados = round((cerrados / total * 100), 1) if total > 0 else 0

    # 2. Datos Semanales (Gráfico de Barras)
    # Obtenemos los últimos 7 días
    hoy = obtener_hora_chile().date()
    fecha_inicio = hoy - timedelta(days=6) # 7 días contando hoy
    
    # Query agrupada por fecha (versión compatible universalmente: traemos datos y procesamos en python)
    # Filtramos por fecha >= inicio Y filtros de seguridad
    weekly_raw = db.session.query(Caso.fecha_ingreso).filter(
        Caso.fecha_ingreso >= fecha_inicio,
        *filters
    ).all()

    # Procesamiento en Python para llenar días vacíos con 0
    weekly_map = {}
    # Inicializar diccionario con los últimos 7 días en 0
    for i in range(7):
        d = fecha_inicio + timedelta(days=i)
        # Guardamos como string 'YYYY-MM-DD' para comparar
        weekly_map[d.strftime('%Y-%m-%d')] = 0

    # Llenar con datos reales
    for row in weekly_raw:
        # Ajuste simple: si la fecha coincide (ignorando hora), sumamos
        # Nota: fecha_ingreso en DB es DateTime. 
        if row.fecha_ingreso:
            d_key = row.fecha_ingreso.date().strftime('%Y-%m-%d')
            if d_key in weekly_map:
                weekly_map[d_key] += 1

    # Preparar listas ordenadas para Chart.js
    nombres_dias = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    bar_labels = []
    bar_data = []

    for i in range(7):
        d = fecha_inicio + timedelta(days=i)
        d_str = d.strftime('%Y-%m-%d')
        weekday_name = nombres_dias[d.weekday()] # 0=Lun, 6=Dom
        
        bar_labels.append(weekday_name) # Ej: "Lun"
        bar_data.append(weekly_map[d_str])

    # 3. Recintos de Notificación (Doughnut) - Agrupar por nombre de recinto
    notif_query = db.session.query(
        CatalogoRecinto.nombre,
        func.count(Caso.id).label('count')
    ).join(Caso.recinto_notifica)\
     .filter(*filters)\
     .group_by(CatalogoRecinto.nombre)\
     .order_by(func.count(Caso.id).desc()).all()

    notif_labels = []
    notif_values = []
    notif_total = 0
    notif_data_full = [] # Para la leyenda HTML

    # Calcular total para porcentajes
    temp_total_notif = sum(row.count for row in notif_query)

    # Definir colores fijos para la leyenda (ciclo de 4 colores)
    colors = ['#3B82F6', '#FBBF24', '#22C55E', '#A855F7', '#EC4899', '#6B7280'] # Azul, Amarillo, Verde, Morado, Rosa, Gris

    for idx, row in enumerate(notif_query):
        pct = round((row.count / temp_total_notif * 100), 1) if temp_total_notif > 0 else 0
        notif_labels.append(row.nombre)
        notif_values.append(row.count)
        
        # Guardamos estructura completa para renderizar leyenda en HTML con Jinja
        notif_data_full.append({
            'nombre': row.nombre,
            'count': row.count,
            'pct': pct,
            'color': colors[idx % len(colors)] # Asignar color cíclico
        })
    
    notif_total = temp_total_notif

    # 4. Recintos Inscritos (Barras Horizontales - Top 5)
    inscritos_query = db.session.query(
        CatalogoEstablecimiento.nombre,
        func.count(Caso.id).label('count')
    ).join(Caso.recinto_inscrito)\
     .filter(*filters)\
     .group_by(CatalogoEstablecimiento.nombre)\
     .order_by(func.count(Caso.id).desc())\
     .limit(5).all()

    inscritos_labels = [row.nombre for row in inscritos_query]
    inscritos_values = [row.count for row in inscritos_query]

    # Empaquetamos todo
    dashboard_data = {
        'total': total,
        'pendientes': pendientes,
        'seguimiento': seguimiento,
        'cerrados': cerrados,
        'pct_pendientes': pct_pendientes,
        'pct_seguimiento': pct_seguimiento,
        'pct_cerrados': pct_cerrados,

        'bar_labels': bar_labels,
        'bar_data': bar_data,

        # Datos Recintos Notificación
        'notif_labels': notif_labels,
        'notif_values': notif_values,
        'notif_total': notif_total,
        'notif_data_full': notif_data_full, # Lista rica para leyenda HTML
        
        # Datos Recintos Inscritos
        'inscritos_labels': inscritos_labels,
        'inscritos_values': inscritos_values
    }

    # =========================================================
    # C. TABLA (CON BÚSQUEDA Y FILTROS)
    # =========================================================
    tabla_query = base_query # Hereda filtros de seguridad

    # Filtro Texto
    if search_query:
        search_clean = search_query.replace('.', '')
        search_pattern = f"%{search_clean}%"
        tabla_query = tabla_query.filter(or_(
            Caso.folio_atencion.ilike(search_pattern),
            Caso.origen_nombres.ilike(search_pattern),
            Caso.origen_apellidos.ilike(search_pattern),
            Caso.paciente_doc_numero.ilike(search_pattern),
            Caso.origen_rut.ilike(search_pattern),
            Caso.acompanante_nombre.ilike(search_pattern)
        ))

    # Filtro Estado
    if estado_filter:
        tabla_query = tabla_query.filter(Caso.estado == estado_filter)

    # Ordenamiento
    orden_estado = case(
        (Caso.estado == 'PENDIENTE_RESCATAR', 0),
        (Caso.estado == 'EN_SEGUIMIENTO', 1),
        (Caso.estado == 'CERRADO', 2),
        else_=3
    )

    pagination = tabla_query.order_by(
        orden_estado, 
        Caso.fecha_ingreso.desc()
    ).paginate(page=page, per_page=15, error_out=False)

    return render_template(
        'casos/index.html',
        pagination=pagination,
        nombre_filtro=titulo_vista,
        stats=dashboard_data,
        candidatos_subrogancia=candidatos_subrogancia,
        subrogante_activo=subrogante_activo
    )

@casos_bp.route('/ver/<int:id>', methods=['GET', 'POST'])
def ver_caso(id):
    caso = Caso.query.get_or_404(id)
    rol_nombre = current_user.rol.nombre

    # --- 1. SEGURIDAD DE ACCESO (Permisos) ---
    permitido = False
    
    if rol_nombre == 'Admin':
        permitido = True
        
    elif rol_nombre in ['Referente', 'Visualizador']:
        # Acceso propio (global o su ciclo)
        acceso_propio = (current_user.ciclo_asignado_id is None) or (caso.ciclo_vital_id == current_user.ciclo_asignado_id)

        # Acceso por subrogancia (si estoy subrogando a un titular con ciclo)
        acceso_subrogado = False
        if current_user.subrogante_de and current_user.subrogante_de.ciclo_asignado_id:
            if caso.ciclo_vital_id == current_user.subrogante_de.ciclo_asignado_id:
                acceso_subrogado = True

        if acceso_propio or acceso_subrogado:
            permitido = True
            
    elif rol_nombre == 'Funcionario':
        # Solo si está asignado a él
        if caso.asignado_a_usuario_id == current_user.id:
            permitido = True
    
    if not permitido:
        abort(403) # Acceso Denegado

    # --- 2. LÓGICA DE ASIGNACIÓN (Solo Admin/Referente) ---
    funcionarios_disponibles = []
    
    # Solo Admin y Referente pueden asignar. Visualizador solo mira. Si el caso está cerrado, no se procesa asignación.
    if rol_nombre in ['Admin', 'Referente'] and caso.estado != 'CERRADO':
        
        # Cargar lista de funcionarios para el select
        q_func = Usuario.query.join(Rol).filter(Rol.nombre == 'Funcionario').filter(Usuario.activo == True)
        
        # Si es Referente de ciclo específico, solo listar funcionarios de ESE ciclo
        if rol_nombre == 'Referente':
            # Mostrar funcionarios del ciclo del CASO (clave para subrogancia)
            q_func = q_func.filter(Usuario.ciclo_asignado_id == caso.ciclo_vital_id)
            
        funcionarios_disponibles = q_func.all()

        # Procesar Formulario de Asignación
        if request.method == 'POST' and 'asignar_funcionario' in request.form:
            # Doble check por si forzaron el POST
            if caso.estado == 'CERRADO':
                flash('El caso está cerrado. No se puede asignar', 'danger')
                return redirect(url_for('casos.ver_caso', id=caso.id))
            
            nuevo_asignado_id = request.form.get('funcionario_id')
            
            if nuevo_asignado_id:
                try:
                    nuevo_asignado_id = int(nuevo_asignado_id)
                    funcionario_nuevo = Usuario.query.get(nuevo_asignado_id)

                    # Validaciones extra de seguridad
                    if not funcionario_nuevo or not funcionario_nuevo.activo:
                        flash('El funcionario seleccionado no es válido o está inactivo.', 'danger')
                        return redirect(url_for('casos.ver_caso', id=caso.id))

                    if rol_nombre == 'Referente':
                        if funcionario_nuevo.ciclo_asignado_id != caso.ciclo_vital_id:
                            flash('El funcionario debe pertenecer al mismo ciclo vital del caso.', 'danger')
                            return redirect(url_for('casos.ver_caso', id=caso.id))
                    
                    # Datos previos para auditoría
                    anterior_asignado_id = caso.asignado_a_usuario_id
                    accion_tipo = 'REASIGNACION' if anterior_asignado_id else 'ASIGNACION'
                    
                    # --- A) ACTUALIZAR CASO Y DB (PRIMERO) ---
                    caso.asignado_a_usuario_id = nuevo_asignado_id
                    caso.asignado_por_usuario_id = current_user.id
                    caso.asignado_at = obtener_hora_chile() # Hora Local
                    
                    # Avanzar estado automáticamente si estaba pendiente
                    if caso.estado == 'PENDIENTE_RESCATAR':
                        caso.estado = 'EN_SEGUIMIENTO'
                    
                    # Auditoría de la Asignación
                    detalles = {
                        'folio': caso.folio_atencion,
                        'previo_id': anterior_asignado_id,
                        'nuevo_id': nuevo_asignado_id,
                        'nombre_asignado': funcionario_nuevo.nombre_completo,
                        'asignado_por': current_user.nombre_completo
                    }
                    
                    nueva_auditoria = AuditoriaCaso(
                        caso_id=caso.id,
                        usuario_id=current_user.id,
                        fecha_movimiento=obtener_hora_chile(),
                        accion=accion_tipo,
                        detalles_cambio=detalles
                    )
                    db.session.add(nueva_auditoria)
                    
                    # COMMIT DE LA ASIGNACIÓN (Esto asegura que el cambio persista sí o sí)
                    db.session.commit()
                    
                    # Log General
                    registrar_log("Asignación Caso", f"Caso #{caso.folio_atencion} asignado a {funcionario_nuevo.nombre_completo}")
                    flash(f'Caso asignado correctamente a {funcionario_nuevo.nombre_completo}', 'success')
                    
                    # --- B) ENVÍO DE CORREO (BEST EFFORT) ---
                    # Lo hacemos DESPUÉS del commit. Si falla, no rompe la asignación.
                    try:
                        email_enviado = enviar_aviso_asignacion(funcionario_nuevo, caso, current_user)
                        
                        if email_enviado:
                            # Log éxito email
                            registrar_log("Email Asignación", f"Enviado a {funcionario_nuevo.email} (Folio: {caso.folio_atencion})")
                            
                            # Auditoría opcional recomendada
                            audit_email = AuditoriaCaso(
                                caso_id=caso.id,
                                usuario_id=current_user.id,
                                fecha_movimiento=obtener_hora_chile(),
                                accion='EMAIL_ASIGNACION',
                                detalles_cambio={'status': 'OK', 'destino': funcionario_nuevo.email}
                            )
                            db.session.add(audit_email)
                            db.session.commit()
                        else:
                            # Log fallo email
                            registrar_log("Error Email", f"Fallo al enviar a {funcionario_nuevo.email}")
                            flash("Aviso: El caso fue asignado, pero no se pudo enviar el correo de notificación.", "warning")
                            
                            audit_email_fail = AuditoriaCaso(
                                caso_id=caso.id,
                                usuario_id=current_user.id,
                                fecha_movimiento=obtener_hora_chile(),
                                accion='EMAIL_ASIGNACION',
                                detalles_cambio={'status': 'ERROR_ENVIO', 'destino': funcionario_nuevo.email}
                            )
                            db.session.add(audit_email_fail)
                            db.session.commit()

                    except Exception as e_mail:
                        print(f"Excepción crítica enviando correo: {e_mail}")
                        registrar_log("Error Crítico Email", str(e_mail))
                    
                    return redirect(url_for('casos.ver_caso', id=caso.id))
                    
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error al asignar caso: {str(e)}', 'danger')

    return render_template('casos/ver.html', 
                           caso=caso, 
                           funcionarios=funcionarios_disponibles)

# --- NUEVA RUTA: GESTIÓN CLÍNICA (FASE 4 P2) ---
@casos_bp.route('/gestionar/<int:id>', methods=['GET', 'POST'])
@login_required
def gestionar_caso(id):
    """
    Formulario para que el Funcionario (o Admin) ingrese la gestión clínica y cierre el caso.
    """
    caso = Caso.query.get_or_404(id)
    rol_nombre = current_user.rol.nombre

    #Si el caso está cerrado, se expulsa inmediatamente
    if caso.estado == 'CERRADO':
        flash('El caso ya está cerrado. No es posible editar la gestión.', 'danger')
        return redirect(url_for('casos.ver_caso', id=caso.id))

    # 1. Validar Permisos para GESTIONAR (Más estricto que ver)
    puede_gestionar = False
    if rol_nombre == 'Admin':
        puede_gestionar = True
    elif rol_nombre == 'Funcionario' and caso.asignado_a_usuario_id == current_user.id:
        puede_gestionar = True
    
    if not puede_gestionar:
        flash('No tienes permisos para gestionar este caso.', 'danger')
        return redirect(url_for('casos.ver_caso', id=caso.id))

    # Cargar catálogos necesarios para el formulario de gestión
    establecimientos = CatalogoEstablecimiento.query.filter_by(activo=True).order_by(CatalogoEstablecimiento.nombre).all()
    instituciones = CatalogoInstitucion.query.filter_by(activo=True).order_by(CatalogoInstitucion.nombre).all()  # Para denuncia

    if request.method == 'POST':
        try:
            # 2. PROCESAR GESTIÓN
            
            # A) Completar Nombres (Solo si venían vacíos del origen)
            # Esto evita sobrescribir si ya tenían datos
            if not caso.origen_nombres:
                caso.origen_nombres = request.form.get('nombres_edit')
            if not caso.origen_apellidos:
                caso.origen_apellidos = request.form.get('apellidos_edit')

            # A2) Completar / Actualizar Datos PACIENTE (opcionales)
            # Nota: Solo rellenamos si viene dato. No sobreescribimos con vacío.
            # Documento
            doc_tipo = request.form.get('paciente_doc_tipo')
            doc_num_raw = request.form.get('paciente_doc_numero')

            if doc_tipo:
                caso.paciente_doc_tipo = doc_tipo

            doc_num_clean = clean(doc_num_raw)
            if doc_num_clean:
                # Si es RUT, normalizamos + validamos (Módulo 11) + largo
                if (doc_tipo == 'RUT'):
                    rut_norm = clean_rut(doc_num_clean)
                    if rut_excede_largo(rut_norm):
                        flash('El RUT del paciente es demasiado largo.', 'danger')
                        return redirect(url_for('casos.gestionar_caso', id=caso.id))
                    if not es_rut_valido(rut_norm):
                        flash('El RUT del paciente no es válido.', 'danger')
                        return redirect(url_for('casos.gestionar_caso', id=caso.id))
                    caso.paciente_doc_numero = rut_norm
                    caso.origen_rut = rut_norm  # Legacy consistente
                else:
                    caso.paciente_doc_numero = doc_num_clean

            # Otro Doc (solo si tipo OTRO y viene descripción)
            if doc_tipo == 'OTRO':
                otro_desc = clean(request.form.get('paciente_doc_otro_desc'))
                if otro_desc:
                    caso.paciente_doc_otro_descripcion = otro_desc

            # Fecha nacimiento (si venía vacía y ahora la completan)
            if not caso.paciente_fecha_nacimiento and request.form.get('paciente_fecha_nacimiento'):
                caso.paciente_fecha_nacimiento = datetime.strptime(request.form.get('paciente_fecha_nacimiento'), '%Y-%m-%d').date()

            # Dirección paciente (si venía vacía y ahora la completan)
            p_calle_new = clean(request.form.get('paciente_calle'))
            p_num_new = clean(request.form.get('paciente_numero'))
            if p_calle_new is not None:
                caso.paciente_direccion_calle = p_calle_new
            if p_num_new is not None:
                caso.paciente_direccion_numero = p_num_new

            # Reconstruir domicilio si tenemos al menos una parte
            p_calle = caso.paciente_direccion_calle
            p_num = caso.paciente_direccion_numero
            caso.paciente_domicilio = f"{p_calle} #{p_num}".strip(" #") if (p_calle or p_num) else None

            # A3) Completar / Actualizar DENUNCIA (permite completar datos faltantes)
            # Si el usuario cambia a "NO", limpiamos todo.
            if request.form.get('denuncia_realizada') is not None:
                denuncia_flag = (request.form.get('denuncia_realizada') == '1')
                caso.denuncia_realizada = denuncia_flag

                if not denuncia_flag:
                    caso.denuncia_institucion_id = None
                    caso.denuncia_institucion_otro = None
                    caso.denuncia_profesional_nombre = None
                    caso.denuncia_profesional_cargo = None
                else:
                    inst_id_int = safe_int(request.form.get('institucion_id'))
                    caso.denuncia_institucion_id = inst_id_int

                    caso.denuncia_institucion_otro = None
                    if inst_id_int:
                        inst_obj = CatalogoInstitucion.query.get(inst_id_int)
                        if inst_obj and 'otro' in inst_obj.nombre.lower():
                            caso.denuncia_institucion_otro = clean(request.form.get('institucion_otro'))

                    nombre_prof = clean(request.form.get('denuncia_nombre'))
                    cargo_prof = clean(request.form.get('denuncia_cargo'))
                    if nombre_prof is not None:
                        caso.denuncia_profesional_nombre = nombre_prof
                    if cargo_prof is not None:
                        caso.denuncia_profesional_cargo = cargo_prof

            # A4) Completar / Actualizar ACOMPAÑANTE (permite completar datos faltantes)
            # Nota: Solo rellenamos si viene dato. No sobreescribimos con vacío.
            acomp_nombre = clean(request.form.get('acomp_nombre'))
            acomp_parentesco = clean(request.form.get('acomp_parentesco'))
            acomp_tel = clean(request.form.get('acomp_telefono'))
            acomp_tel_tipo = request.form.get('acomp_tel_tipo')

            if acomp_nombre is not None:
                caso.acompanante_nombre = acomp_nombre
            if acomp_parentesco is not None:
                caso.acompanante_parentesco = acomp_parentesco
            if acomp_tel is not None:
                caso.acompanante_telefono = acomp_tel
            if acomp_tel_tipo:
                caso.acompanante_telefono_tipo = acomp_tel_tipo

            acomp_doc_tipo = request.form.get('acomp_doc_tipo')
            if acomp_doc_tipo:
                caso.acompanante_doc_tipo = acomp_doc_tipo

            acomp_doc_num_raw = request.form.get('acomp_doc_numero')
            acomp_doc_num_clean = clean(acomp_doc_num_raw)
            if acomp_doc_num_clean:
                if acomp_doc_tipo == 'RUT':
                    rut_acomp_norm = clean_rut(acomp_doc_num_clean)
                    if rut_excede_largo(rut_acomp_norm):
                        flash('El RUT del acompañante es demasiado largo.', 'danger')
                        return redirect(url_for('casos.gestionar_caso', id=caso.id))
                    if not es_rut_valido(rut_acomp_norm):
                        flash('El RUT del acompañante no es válido.', 'danger')
                        return redirect(url_for('casos.gestionar_caso', id=caso.id))
                    caso.acompanante_doc_numero = rut_acomp_norm
                else:
                    caso.acompanante_doc_numero = acomp_doc_num_clean

            if acomp_doc_tipo == 'OTRO':
                acomp_otro_desc = clean(request.form.get('acomp_doc_otro_desc'))
                if acomp_otro_desc:
                    caso.acompanante_doc_otro_descripcion = acomp_otro_desc

            a_calle_new = clean(request.form.get('acomp_calle'))
            a_num_new = clean(request.form.get('acomp_numero'))
            if a_calle_new is not None:
                caso.acompanante_direccion_calle = a_calle_new
            if a_num_new is not None:
                caso.acompanante_direccion_numero = a_num_new

            a_calle = caso.acompanante_direccion_calle
            a_num = caso.acompanante_direccion_numero
            caso.acompanante_domicilio = f"{a_calle} #{a_num}".strip(" #") if (a_calle or a_num) else None

            # B) Datos Clínicos (CORRECCIÓN SAFE INT + OTRO)
            recinto_inscrito_id_raw = request.form.get('recinto_inscrito_id')
            recinto_inscrito_id_int = safe_int(recinto_inscrito_id_raw) # Convierte '' a None
            
            caso.recinto_inscrito_id = recinto_inscrito_id_int
            
            # Lógica "Otro Recinto Inscrito"
            caso.recinto_inscrito_otro_texto = None # Reset por defecto
            if recinto_inscrito_id_int:
                est_obj = CatalogoEstablecimiento.query.get(recinto_inscrito_id_int)
                if est_obj and 'otro' in est_obj.nombre.lower():
                    texto_otro = request.form.get('recinto_inscrito_otro')
                    if texto_otro and texto_otro.strip():
                        caso.recinto_inscrito_otro_texto = texto_otro.strip()
                    else:
                        flash("Seleccionó 'Otro' recinto pero no especificó cuál.", "warning")
                        # No bloqueamos, pero avisamos.

            caso.ingreso_lain = (request.form.get('ingreso_lain') == '1')
            
            # C) Lógica Fallecido
            es_fallecido = (request.form.get('fallecido') == '1')
            caso.fallecido = es_fallecido
            if es_fallecido and request.form.get('fecha_defuncion'):
                caso.fecha_defuncion = datetime.strptime(request.form.get('fecha_defuncion'), '%Y-%m-%d').date()
            else:
                caso.fecha_defuncion = None # Limpiar si desmarcan

            # D) Seguimiento (Los valores vienen exactamente como los códigos ENUM del select)
            caso.control_sanitario = request.form.get('control_sanitario')
            caso.gestion_vacunas = request.form.get('gestion_vacunas')
            caso.gestion_judicial = request.form.get('gestion_judicial')
            caso.gestion_salud_mental = request.form.get('gestion_salud_mental')
            caso.gestion_cosam = request.form.get('gestion_cosam')

            # E) Observaciones Finales (NUEVA LÓGICA: BITÁCORA)
            # Obtenemos el texto del textarea nuevo
            nueva_obs = request.form.get('nueva_observacion')
            
            if nueva_obs and nueva_obs.strip():
                # Crear registro en la bitácora clínica
                nueva_gestion = CasoGestion(
                    caso_id=caso.id,
                    usuario_id=current_user.id,
                    fecha_movimiento=obtener_hora_chile(),
                    observacion=nueva_obs.strip()
                )
                db.session.add(nueva_gestion)
                
                # Opcional: Actualizamos el campo legacy solo para tener "lo último" a mano rápido,
                # pero la verdad reside en la tabla nueva.
                caso.observaciones_gestion = nueva_obs.strip()
            
            # F) Auditoría del movimiento
            audit = AuditoriaCaso(
                caso_id=caso.id,
                usuario_id=current_user.id,
                fecha_movimiento=obtener_hora_chile(),
                accion='GESTION_CLINICA',
                detalles_cambio={'tipo': 'actualizacion_seguimiento'}
            )
            db.session.add(audit)
            
            db.session.commit()
            
            flash('Gestión guardada exitosamente.', 'success')
            return redirect(url_for('casos.ver_caso', id=caso.id))

        except Exception as e:
            db.session.rollback()
            print(f"Error gestionando caso: {e}")
            flash('Ocurrió un error al guardar la gestión.', 'danger')

    return render_template('casos/gestion.html', caso=caso, establecimientos=establecimientos, instituciones=instituciones)

# --- RUTA CIERRE DE CASO (FINAL) ---
@casos_bp.route('/cerrar/<int:id>', methods=['POST'])
@login_required
def cerrar_caso(id):
    caso = Caso.query.get_or_404(id)
    rol_nombre = current_user.rol.nombre

    # 1. Validar Permisos (Admin o Funcionario Asignado)
    puede_cerrar = False
    if rol_nombre == 'Admin':
        puede_cerrar = True
    elif rol_nombre == 'Funcionario' and caso.asignado_a_usuario_id == current_user.id:
        puede_cerrar = True
    
    if not puede_cerrar:
        flash('No tienes permisos para cerrar este caso.', 'danger')
        return redirect(url_for('casos.ver_caso', id=caso.id))

    if caso.estado == 'CERRADO':
        flash('El caso ya se encuentra cerrado.', 'warning')
        return redirect(url_for('casos.ver_caso', id=caso.id))

    try:
        # 2. Actualizar Estado en BD (Primer Commit)
        caso.estado = 'CERRADO'
        caso.fecha_cierre = obtener_hora_chile() # Fecha oficial
        caso.usuario_cierre_id = current_user.id
        
        # Auditoría
        audit = AuditoriaCaso(
            caso_id=caso.id,
            usuario_id=current_user.id,
            fecha_movimiento=obtener_hora_chile(),
            accion='CIERRE_CASO',
            detalles_cambio={'motivo': 'Cierre manual por gestión finalizada'}
        )
        db.session.add(audit)
        db.session.commit() # Guardamos para que la fecha_cierre esté firme en DB

        # 3. Generar PDF de Acta (Path Estable y Multiplataforma)
        filename = f"acta_{caso.id}_{caso.folio_atencion}.pdf"
        
        # Calculamos la raíz del proyecto subiendo un nivel desde 'blueprints/'
        # Esto funciona en Windows y Linux/cPanel por igual
        BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        uploads_actas_dir = os.path.join(BASE_DIR, 'uploads', 'actas')
        
        # Ruta absoluta completa para guardar el archivo
        output_path_abs = os.path.join(uploads_actas_dir, filename)

        # Ruta relativa para guardar en BD (portable)
        output_path_rel = f"uploads/actas/{filename}"

        # Generar el PDF
        generar_acta_cierre_pdf(caso, output_path_abs, current_user)

        # 4. Guardar ruta relativa en BD
        caso.acta_pdf_path = output_path_rel
        db.session.commit()

        # 5. Enviar Correo con Adjunto (Best Effort)
        try:
            enviar_aviso_cierre(caso, current_user)
            flash('Caso cerrado exitosamente. Acta generada y notificaciones enviadas.', 'success')
        except Exception as e_mail:
            print(f"Error enviando correo cierre: {e_mail}")
            flash('Caso cerrado y acta generada, pero falló el envío del correo.', 'warning')

        registrar_log("Cierre Caso", f"Caso #{caso.folio_atencion} cerrado por {current_user.nombre_completo}")

    except Exception as e:
        db.session.rollback()
        print(f"Error cerrando caso: {e}")
        flash(f'Error crítico al cerrar el caso: {str(e)}', 'danger')

    return redirect(url_for('casos.ver_caso', id=caso.id))

# --- RUTA DESCARGA SEGURA DE ACTA (PRODUCCIÓN) ---
@casos_bp.route('/acta/<int:id>', methods=['GET'])
@login_required
def descargar_acta(id):
    """
    Endpoint protegido para descargar el PDF del acta.
    Incluye protección robusta contra Path Traversal, auditoría de descarga
    y validación de permisos extendida (asignado o cerrador).
    """
    caso = Caso.query.get_or_404(id)
    rol_nombre = current_user.rol.nombre

    # 1. VALIDACIÓN DE PERMISOS
    permitido = False
    
    if rol_nombre == 'Admin':
        permitido = True
        
    elif rol_nombre in ['Referente', 'Visualizador']:
        acceso_propio = (current_user.ciclo_asignado_id is None) or (caso.ciclo_vital_id == current_user.ciclo_asignado_id)

        acceso_subrogado = False
        if current_user.subrogante_de and current_user.subrogante_de.ciclo_asignado_id:
            if caso.ciclo_vital_id == current_user.subrogante_de.ciclo_asignado_id:
                acceso_subrogado = True

        if acceso_propio or acceso_subrogado:
            permitido = True
            
    elif rol_nombre == 'Funcionario':
        # AJUSTE PERMISOS: Permitido si es el asignado actual O quien cerró el caso
        es_asignado = (caso.asignado_a_usuario_id == current_user.id)
        es_quien_cerro = (caso.usuario_cierre_id == current_user.id)
        
        if es_asignado or es_quien_cerro:
            permitido = True
            
    if not permitido:
        flash("No tienes permisos para descargar este documento.", "danger")
        return redirect(url_for('casos.index'))

    # 2. VALIDAR EXISTENCIA EN BD
    if not caso.acta_pdf_path:
        flash('El caso no tiene un acta generada.', 'warning')
        return redirect(url_for('casos.ver_caso', id=caso.id))

    try:
        # 3. CONSTRUCCIÓN Y SEGURIDAD DE RUTA (PATH TRAVERSAL ROBUSTO)
        
        # Base del proyecto (un nivel arriba de blueprints/)
        BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        
        # Limpiamos la ruta relativa de BD
        ruta_relativa = caso.acta_pdf_path.replace('\\', '/').lstrip('/')
        
        # Resolvemos la ruta absoluta final del archivo solicitado
        path_absoluto = os.path.abspath(os.path.join(BASE_DIR, ruta_relativa))
        
        # Definimos la "jaula" autorizada (carpeta actas)
        carpeta_actas = os.path.abspath(os.path.join(BASE_DIR, 'uploads', 'actas'))

        # AJUSTE 1: Validación estricta con os.sep para evitar falsos positivos
        # Aseguramos que la ruta comience con "carpeta_actas/" (o "\" en Windows)
        # Esto evita que 'uploads/actas_secretas' pase el filtro de 'uploads/actas'
        carpeta_con_sep = os.path.join(carpeta_actas, '') # Agrega el separador al final (/ o \)
        
        if not path_absoluto.startswith(carpeta_con_sep):
            registrar_log("Seguridad", f"ALERTA: Intento de Path Traversal por {current_user.email}. Path: {path_absoluto}")
            abort(403) # Acceso Prohibido

        # 4. VERIFICAR ARCHIVO FÍSICO
        if not os.path.exists(path_absoluto):
            registrar_log("Error Archivo", f"Acta no encontrada en disco: {path_absoluto}")
            flash('El archivo físico del acta no se encuentra en el servidor.', 'danger')
            return redirect(url_for('casos.ver_caso', id=caso.id))

        # AJUSTE 2: Auditoría de la descarga (Trazabilidad)
        # Registramos quién descargó el archivo y de qué caso antes de enviarlo
        registrar_log(
            "Descarga Acta", 
            f"Usuario={current_user.email} descargó el acta del Caso={caso.id} - Folio={caso.folio_atencion}"
        )

        # AJUSTE 3: Nombre de archivo amigable para el usuario
        # Genera "Acta_Cierre_FOLIO.pdf" en lugar de "acta_1_FOLIO.pdf" (interno)
        nombre_descarga = f"Acta_Cierre_{caso.folio_atencion or caso.id}.pdf"

        # 5. SERVIR ARCHIVO
        return send_file(
            path_absoluto,
            as_attachment=True,
            download_name=nombre_descarga
        )

    except Exception as e:
        print(f"Error sirviendo archivo: {e}")
        flash('Error interno al intentar descargar el archivo.', 'danger')
        return redirect(url_for('casos.ver_caso', id=caso.id))
    
# --- NUEVA RUTA: EXPORTAR EXCEL ---
@casos_bp.route('/exportar')
@login_required
def exportar_excel():
    """
    Genera y descarga un reporte Excel (.xlsx) de los casos filtrados.
    Reutiliza la MISMA lógica de seguridad y filtros que el index.
    """
    # 1. Recuperar filtros de la URL (igual que en index)
    search_query = request.args.get('search', '').strip()
    estado_filter = request.args.get('estado', '').strip()

    rol_nombre = current_user.rol.nombre
    
    # 2. Construir Filtros de Seguridad (Base Query)
    filters = []

    if rol_nombre == 'Admin':
        pass
    elif rol_nombre in ['Referente', 'Visualizador']:
        ciclos_permitidos = []

        if current_user.ciclo_asignado_id:
            ciclos_permitidos.append(current_user.ciclo_asignado_id)

        if current_user.subrogante_de and current_user.subrogante_de.ciclo_asignado_id:
            ciclo_sub = current_user.subrogante_de.ciclo_asignado_id
            if ciclo_sub not in ciclos_permitidos:
                ciclos_permitidos.append(ciclo_sub)

        if ciclos_permitidos:
            filters.append(Caso.ciclo_vital_id.in_(ciclos_permitidos))
        else:
            # sin ciclo = global (no filtra)
            pass
    elif rol_nombre == 'Funcionario':
        filters.append(Caso.asignado_a_usuario_id == current_user.id)
    else:
        abort(403)

    # Iniciar query segura
    query = Caso.query.filter(*filters)

    # 3. Aplicar Filtros de Usuario (Búsqueda y Estado)
    if search_query:
        search_clean = search_query.replace('.', '')
        search_pattern = f"%{search_clean}%"
        query = query.filter(or_(
            Caso.folio_atencion.ilike(search_pattern),
            Caso.origen_nombres.ilike(search_pattern),
            Caso.origen_apellidos.ilike(search_pattern),
            Caso.paciente_doc_numero.ilike(search_pattern),
            Caso.origen_rut.ilike(search_pattern),
            Caso.acompanante_nombre.ilike(search_pattern)
        ))

    if estado_filter:
        query = query.filter(Caso.estado == estado_filter)

    # 4. Obtener TODOS los resultados (Sin paginación)
    # Ordenamos igual que la vista: Prioridad Estado -> Fecha
    orden_estado = case(
        (Caso.estado == 'PENDIENTE_RESCATAR', 0),
        (Caso.estado == 'EN_SEGUIMIENTO', 1),
        (Caso.estado == 'CERRADO', 2),
        else_=3
    )
    casos = query.order_by(orden_estado, Caso.fecha_ingreso.desc()).all()

    # 5. Generar Excel con openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Casos"

    # Definir Encabezados
    headers = [
        "Fecha Ingreso", 
        "Folio", 
        "Estado", 
        "Tipo Doc", 
        "Num Documento", 
        "Paciente", 
        "Edad (Ref)",
        "Ciclo Vital", 
        "Recinto Notifica", 
        "Recinto Inscrito",
        "Asignado A",
        "Fecha Cierre"
    ]

    # Estilo para Encabezados (Negrita, Fondo Azul Oscuro, Letra Blanca)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Gris oscuro tipo Tailwind gray-800
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Escribir Encabezados
    ws.append(headers)
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center

    # Escribir Datos
    for caso in casos:
        # Pre-procesamiento de datos para evitar errores con None
        fecha_ingreso = caso.fecha_ingreso.strftime('%d-%m-%Y %H:%M') if caso.fecha_ingreso else ""
        fecha_cierre = caso.fecha_cierre.strftime('%d-%m-%Y %H:%M') if caso.fecha_cierre else ""
        
        # Nombre completo paciente
        nombres = caso.origen_nombres or ""
        apellidos = caso.origen_apellidos or ""
        paciente_full = f"{nombres} {apellidos}".strip()

        # Documento
        doc_tipo = caso.paciente_doc_tipo or "RUT"
        doc_num = caso.paciente_doc_numero or caso.origen_rut or "S/I"

        # Asignado
        asignado = caso.asignado_a.nombre_completo if caso.asignado_a else "Sin Asignar"

        # Recintos
        recinto_notifica = caso.recinto_notifica.nombre if caso.recinto_notifica else "Desconocido"
        recinto_inscrito = caso.recinto_inscrito.nombre if caso.recinto_inscrito else "No Registrado"
        if caso.recinto_inscrito_otro_texto:
            recinto_inscrito += f" ({caso.recinto_inscrito_otro_texto})"

        # Edad referencial (calculo simple si hay fecha nac)
        edad_str = ""
        if caso.paciente_fecha_nacimiento:
            hoy = obtener_hora_chile().date()
            nac = caso.paciente_fecha_nacimiento
            edad = hoy.year - nac.year - ((hoy.month, hoy.day) < (nac.month, nac.day))
            edad_str = f"{edad} años"

        row = [
            fecha_ingreso,
            caso.folio_atencion,
            caso.estado,
            doc_tipo,
            doc_num,
            paciente_full,
            edad_str,
            caso.ciclo_vital.nombre,
            recinto_notifica,
            recinto_inscrito,
            asignado,
            fecha_cierre
        ]
        ws.append(row)

    # 6. Formato Final (AutoFilter, Congelar Panel, Anchos)
    
    # Agregar AutoFilter
    ws.auto_filter.ref = ws.dimensions

    # Congelar primera fila
    ws.freeze_panes = "A2"

    # Ajustar ancho de columnas (Estimación decente)
    column_widths = [18, 10, 15, 10, 15, 30, 10, 15, 30, 30, 20, 18]
    for i, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # 7. Guardar en memoria y enviar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Reporte_Casos_{obtener_hora_chile().strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@casos_bp.route('/enviar_reporte_masivo', methods=['POST'])
@login_required
def enviar_reporte_masivo():
    """
    Calcula estadísticas globales y las envía por correo a todos los usuarios activos.
    Solo para Admin y Referentes.
    Incluye tablas:
    - Resumen por Recinto (Inscritos) con estados
    - Resumen de Notificaciones (Origen) con % del total
    """
    # 1) Validar permiso
    if current_user.rol.nombre not in ['Admin', 'Referente']:
        flash("No tiene permisos para realizar esta acción.", "danger")
        return redirect(url_for('casos.index'))

    try:
        # 2) Calcular estadísticas globales (snapshot)
        stats_query = db.session.query(
            func.count(Caso.id).label('total'),
            func.sum(case((Caso.estado == 'PENDIENTE_RESCATAR', 1), else_=0)).label('pendientes'),
            func.sum(case((Caso.estado == 'EN_SEGUIMIENTO', 1), else_=0)).label('seguimiento'),
            func.sum(case((Caso.estado == 'CERRADO', 1), else_=0)).label('cerrados')
        )

        r = stats_query.first()
        if not r:
            flash("No fue posible calcular estadísticas del sistema.", "danger")
            return redirect(url_for('casos.index'))

        stats = {
            'total': int(r.total or 0),
            'pendientes': int(r.pendientes or 0),
            'seguimiento': int(r.seguimiento or 0),
            'cerrados': int(r.cerrados or 0)
        }

        # 3) Obtener destinatarios (usuarios activos con email válido)
        usuarios_activos = Usuario.query.filter(
            Usuario.activo == True,
            Usuario.email.isnot(None),
            Usuario.email != ''
        ).all()

        # Limpieza + deduplicación
        destinatarios_bcc = []
        vistos = set()
        for u in usuarios_activos:
            email = (u.email or "").strip().lower()
            if email and email not in vistos:
                vistos.add(email)
                destinatarios_bcc.append(email)

        if not destinatarios_bcc:
            flash("No se encontraron usuarios activos con correo para enviar el reporte.", "warning")
            return redirect(url_for('casos.index'))

        # =========================================================
        # 4) NUEVO: Resumen por Recinto (Inscritos) con estados
        # Incluye "No Registrado" si recinto_inscrito_id es NULL
        # =========================================================
        q_inscritos = db.session.query(
            func.coalesce(CatalogoEstablecimiento.nombre, 'No Registrado').label('nombre'),
            func.count(Caso.id).label('total'),
            func.sum(case((Caso.estado == 'PENDIENTE_RESCATAR', 1), else_=0)).label('pendientes'),
            func.sum(case((Caso.estado == 'EN_SEGUIMIENTO', 1), else_=0)).label('seguimiento'),
            func.sum(case((Caso.estado == 'CERRADO', 1), else_=0)).label('cerrados')
        ).outerjoin(Caso.recinto_inscrito) \
         .group_by(func.coalesce(CatalogoEstablecimiento.nombre, 'No Registrado')) \
         .order_by(func.count(Caso.id).desc()) \
         .all()

        stats_inscritos = []
        for row in q_inscritos:
            stats_inscritos.append({
                'nombre': row.nombre,
                'total': int(row.total or 0),
                'pendientes': int(row.pendientes or 0),
                'seguimiento': int(row.seguimiento or 0),
                'cerrados': int(row.cerrados or 0)
            })

        # =========================================================
        # 5) NUEVO: Resumen de Notificaciones (Origen) + porcentaje
        # Incluye "No especificado" si recinto_notifica_id es NULL
        # =========================================================
        q_notif = db.session.query(
            func.coalesce(CatalogoRecinto.nombre, 'No especificado').label('nombre'),
            func.count(Caso.id).label('total')
        ).outerjoin(Caso.recinto_notifica) \
         .group_by(func.coalesce(CatalogoRecinto.nombre, 'No especificado')) \
         .order_by(func.count(Caso.id).desc()) \
         .all()

        total_notif_global = sum(int(row.total or 0) for row in q_notif) or 0

        stats_notificacion = []
        for row in q_notif:
            total_row = int(row.total or 0)
            pct = round((total_row / total_notif_global * 100), 1) if total_notif_global > 0 else 0
            stats_notificacion.append({
                'nombre': row.nombre,
                'total': total_row,
                'pct': pct
            })

        # 6) Empaquetar todo para email.py (nueva firma)
        data_completa = {
            'global': stats,
            'inscritos': stats_inscritos,
            'notificacion': stats_notificacion
        }

        # 7) Enviar correo masivo
        if enviar_reporte_estadistico_masivo(destinatarios_bcc, data_completa):
            registrar_log("Reporte Masivo", f"Enviado por {current_user.email} a {len(destinatarios_bcc)} destinatarios.")
            flash(f"Reporte enviado exitosamente a {len(destinatarios_bcc)} usuarios.", "success")
        else:
            flash("Hubo un error al intentar enviar el reporte por correo.", "danger")

    except Exception as e:
        print(f"Error reporte masivo: {e}")
        flash(f"Error interno: {str(e)}", "danger")

    return redirect(url_for('casos.index'))

@casos_bp.route('/subrogancia/gestionar', methods=['POST'])
@login_required
def gestionar_subrogancia():
    """
    Activa o desactiva subrogancia para el usuario actual (Titular Referente).
    - Activar: el subrogante apunta al titular: subrogante.subrogante_de_usuario_id = current_user.id
    - Desactivar: rompe vínculo del subrogante activo (si existe)
    """
    if current_user.rol.nombre != 'Referente':
        flash("Solo los Referentes pueden gestionar subrogancias.", "danger")
        return redirect(url_for('casos.index'))

    accion = (request.form.get('accion') or '').strip()

    try:
        # Helper: obtener subrogante activo (compatible con lazy='dynamic' y con listas)
        def get_subrogante_activo():
            rel = getattr(current_user, "subrogantes_activos", None)
            if not rel:
                return None
            try:
                # Si es AppenderQuery (dynamic)
                return rel.first()
            except Exception:
                # Si es lista normal
                return rel[0] if rel else None

        if accion == 'desactivar':
            subrogante_actual = get_subrogante_activo()

            if not subrogante_actual:
                flash("No tienes subrogancia activa para desactivar.", "warning")
                return redirect(url_for('casos.index'))

            subrogante_actual.subrogante_de_usuario_id = None
            db.session.commit()

            # Correo best-effort
            try:
                enviar_aviso_subrogancia(current_user, subrogante_actual, es_activacion=False)
            except Exception as e_mail:
                print(f"Error email subrogancia (desactivar): {e_mail}")

            registrar_log("Subrogancia", f"{current_user.email} desactivó subrogancia de {subrogante_actual.email}")
            flash(f"Subrogancia finalizada. {subrogante_actual.nombre_completo} ya no tiene acceso a tu ciclo.", "info")
            return redirect(url_for('casos.index'))

        if accion == 'activar':
            subrogante_id = safe_int(request.form.get('subrogante_id'))
            if not subrogante_id:
                flash("Debes seleccionar un subrogante.", "warning")
                return redirect(url_for('casos.index'))

            if subrogante_id == current_user.id:
                flash("No puedes autodesignarte como subrogante.", "danger")
                return redirect(url_for('casos.index'))

            subrogante_nuevo = Usuario.query.get(subrogante_id)
            if (not subrogante_nuevo) or (not subrogante_nuevo.activo) or (subrogante_nuevo.rol.nombre != 'Referente'):
                flash("El usuario seleccionado no es válido o no es Referente.", "danger")
                return redirect(url_for('casos.index'))

            # Regla simple: 1 subrogancia activa por titular
            # Si ya existe alguien subrogándome, lo cortamos primero
            anterior = get_subrogante_activo()
            if anterior:
                anterior.subrogante_de_usuario_id = None

            # Activar (el subrogante apunta al titular)
            subrogante_nuevo.subrogante_de_usuario_id = current_user.id
            db.session.commit()

            # Correo best-effort
            try:
                enviar_aviso_subrogancia(current_user, subrogante_nuevo, es_activacion=True)
            except Exception as e_mail:
                print(f"Error email subrogancia (activar): {e_mail}")

            registrar_log("Subrogancia", f"{current_user.email} activó subrogancia a {subrogante_nuevo.email}")
            flash(f"Subrogancia activada. {subrogante_nuevo.nombre_completo} ahora tiene acceso a tu ciclo.", "success")
            return redirect(url_for('casos.index'))

        flash("Acción de subrogancia inválida.", "warning")
        return redirect(url_for('casos.index'))

    except Exception as e:
        db.session.rollback()
        print(f"Error gestionar_subrogancia: {e}")
        flash("Ocurrió un error al gestionar la subrogancia.", "danger")
        return redirect(url_for('casos.index'))